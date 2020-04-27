import openpyxl
import numpy

def loadexcel():
    print("load execel train.xlsx")
    wb = openpyxl.load_workbook("train.xlsx")
    trainsht = wb.get_sheet_by_name("train")
    maxrow = trainsht.max_row
    maxcol = trainsht.max_column
    data = dict()
    trainset = [[]]
    ylist = []
    index = []
    rowindex = 0
    prePM = 26
    index.append("Hour")
    for row_ in range(2,maxrow+1):

        fname = trainsht.cell(row = row_, column = 3 ).value
        if rowindex < 1:
            if fname == "PM2.5":
                index.append("Pre PM2.5")
            else:
                index.append(fname)

        for col_ in range (4, maxcol+1):
            tup_ = []
            if row_ -2 == 0 or ((row_ -2) % 18 ) == 0:

                tup_.append(col_ - 4)
                tup_.append(trainsht.cell(row = row_, column = col_ ).value)
                trainset.append(tup_)
            else:
                if fname == "PM2.5":
                    #trainset[rowindex * 24 + (col_ - 3)].append("prePM:" +str(prePM)+ "from row:" + str(row_) +"col:"+ str(col_) )
                    trainset[rowindex * 24 + (col_ - 3)].append(prePM)
                    curPM = trainsht.cell(row=row_, column=col_).value
                    #ylist.append("curPM:" + str(curPM) + "from row:" + str(row_) +"col:"+ str(col_))
                    ylist.append(curPM)
                    prePM = curPM
                else:
                    if fname == "RAINFALL":
                        rainfall = trainsht.cell(row=row_, column=col_).value
                        if rainfall == "NR":
                            trainset[rowindex * 24 + (col_ - 3)].append(0.0)
                        else:
                            trainset[rowindex * 24 + (col_ - 3)].append(trainsht.cell(row=row_, column=col_).value)
                    else:
                        trainset[rowindex * 24 + (col_ - 3)].append(trainsht.cell(row=row_, column=col_).value)
        if row_ !=2 and ((row_ - 1) % 18) == 0:
            rowindex = rowindex + 1
    del trainset[0]
    data["trainset"] =  numpy.transpose( trainset)
    data["y"] = numpy.transpose(ylist)
    data["index"] = numpy.transpose(index)

    #data["trainset"] = trainset
    #data["y"] = ylist
    #data["index"] = index
    return data







